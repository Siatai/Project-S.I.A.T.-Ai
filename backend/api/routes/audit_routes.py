from __future__ import annotations

import asyncio
import json
import os
import re
import sqlite3
import uuid
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import httpx
from bs4 import BeautifulSoup
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook


router = APIRouter()

PROJECT_DIR = Path(__file__).resolve().parents[3]
AUDIT_DIR = PROJECT_DIR / "audit_service"
AUDIT_DATA_DIR = AUDIT_DIR / "data"
AUDIT_DB_FILE = AUDIT_DATA_DIR / "checker.sqlite"

TRANSFER_TOPIC = "0xddf252ad1be2c89b69c2b068fc378daa952ba7f163c4a11628f55a4df523b3ef"
TRANSFER_SELECTOR = "0xa9059cbb"
audit_sync_lock = asyncio.Lock()
audit_sync_task: asyncio.Task | None = None

CONFIG = {
    "deposit_wallet_address": os.getenv("DEPOSIT_WALLET_ADDRESS", "0xc2b65f40b8361F9eCf27FB03F2ce3992D1F0211c").lower(),
    "withdrawal_wallet_address": os.getenv("WITHDRAWAL_WALLET_ADDRESS", "0x876F2A2EfE1B20E38018c8292823d814bf195216").lower(),
    "usdt_contract": os.getenv("USDT_CONTRACT", "0x55d398326f99059fF775485246999027B3197955").lower(),
    "bsc_api_key": os.getenv("BSCSCAN_API_KEY", ""),
    "bsc_api_base": os.getenv("BSCSCAN_API_BASE", "https://api.bscscan.com/api"),
    "rpc_url": os.getenv("BSC_RPC_URL", "https://public-bsc-mainnet.fastnode.io"),
    "rpc_fallback_urls": [
        item.strip()
        for item in os.getenv("BSC_RPC_FALLBACK_URLS", "https://bsc.leorpc.com/?api_key=FREE").split(",")
        if item.strip()
    ],
    "poll_interval_ms": int(os.getenv("POLL_INTERVAL_MS", "45000")),
    "telegram_bot_token": os.getenv("TELEGRAM_BOT_TOKEN", ""),
    "telegram_chat_id": os.getenv("TELEGRAM_CHAT_ID", ""),
    "match_window_minutes": int(os.getenv("MATCH_WINDOW_MINUTES", "5")),
    "amount_tolerance": float(os.getenv("AMOUNT_TOLERANCE", "1")),
    "initial_backfill_hours": int(os.getenv("INITIAL_BACKFILL_HOURS", "24")),
    "min_transaction_amount": float(os.getenv("MIN_TRANSACTION_AMOUNT", "1")),
}


def ensure_data_dir() -> None:
    AUDIT_DATA_DIR.mkdir(parents=True, exist_ok=True)


def connect_db() -> sqlite3.Connection:
    ensure_data_dir()
    connection = sqlite3.connect(AUDIT_DB_FILE)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode = WAL")
    connection.executescript(
        """
        CREATE TABLE IF NOT EXISTS sync_state (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS transactions (
          tx_key TEXT PRIMARY KEY,
          hash TEXT,
          wallet_address TEXT,
          block_number INTEGER,
          sender TEXT,
          recipient TEXT,
          token_symbol TEXT,
          token_name TEXT,
          contract_address TEXT,
          amount REAL,
          timestamp TEXT,
          raw_json TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS telegram_messages (
          id TEXT PRIMARY KEY,
          tx_hash TEXT,
          fees_tx_hash TEXT,
          message_id INTEGER,
          chat_id TEXT,
          sender TEXT,
          amount REAL,
          timestamp TEXT,
          text TEXT,
          raw_json TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_telegram_messages_tx_hash ON telegram_messages(tx_hash);
        CREATE INDEX IF NOT EXISTS idx_telegram_messages_fees_tx_hash ON telegram_messages(fees_tx_hash);
        CREATE INDEX IF NOT EXISTS idx_telegram_messages_chat_message ON telegram_messages(chat_id, message_id);
        CREATE INDEX IF NOT EXISTS idx_transactions_timestamp ON transactions(timestamp);
        CREATE INDEX IF NOT EXISTS idx_transactions_wallet_address ON transactions(wallet_address);
        CREATE INDEX IF NOT EXISTS idx_transactions_hash ON transactions(hash);
        CREATE INDEX IF NOT EXISTS idx_telegram_messages_timestamp ON telegram_messages(timestamp);
        """
    )
    return connection


def create_wallet_sync_state() -> dict[str, Any]:
    return {
        "trackedAddress": None,
        "lastCheckedAt": None,
        "lastBlock": 0,
        "lastError": None,
        "lastSource": None,
        "baselineTimestamp": None,
        "initialized": False,
        "lastMode": None,
        "lastFullSyncAt": None,
        "lastIncrementalSyncAt": None,
    }


def default_store() -> dict[str, Any]:
    return {
        "transactions": [],
        "telegramMessages": [],
        "manualReconciliations": [],
        "sync": {
            "wallets": {
                "deposit": create_wallet_sync_state(),
                "withdrawal": create_wallet_sync_state(),
            },
            "telegram": {
                "trackedAddress": None,
                "lastCheckedAt": None,
                "lastError": None,
                "lastMode": "live-only",
                "lastMessageAt": None,
                "historicalImportSupported": False,
                "importCursors": {},
            },
        },
    }


def get_wallet_address(key: str) -> str:
    if key == "deposit":
        return CONFIG["deposit_wallet_address"]
    return CONFIG["withdrawal_wallet_address"]


def ensure_wallet_sync_state(sync: dict[str, Any], wallet_key: str) -> dict[str, Any]:
    sync.setdefault("wallets", {})
    sync["wallets"][wallet_key] = {
        **create_wallet_sync_state(),
        **sync["wallets"].get(wallet_key, {}),
    }
    return sync["wallets"][wallet_key]


def build_transaction_key(tx: dict[str, Any]) -> str:
    wallet_address = str(tx.get("walletAddress") or "").lower()
    tx_hash = str(tx.get("hash") or "").lower()
    log_index = tx.get("raw", {}).get("logIndex", tx.get("logIndex"))
    if log_index not in (None, ""):
        return f"{wallet_address}:{tx_hash}:{log_index}"
    return ":".join(
        [
            wallet_address,
            tx_hash,
            str(tx.get("from") or "").lower(),
            str(tx.get("to") or "").lower(),
            f"{float(tx.get('amount') or 0):.8f}",
            str(tx.get("timestamp") or ""),
        ]
    )


def read_store() -> dict[str, Any]:
    store = default_store()
    with connect_db() as connection:
        sync_row = connection.execute("SELECT value FROM sync_state WHERE key = ?", ("store",)).fetchone()
        manual_row = connection.execute("SELECT value FROM sync_state WHERE key = ?", ("manual_reconciliations",)).fetchone()
        if sync_row and sync_row["value"]:
            parsed = json.loads(sync_row["value"])
            store["sync"] = {
                "wallets": {
                    "deposit": {**store["sync"]["wallets"]["deposit"], **(parsed.get("wallets", {}).get("deposit", parsed.get("wallet", {})))},
                    "withdrawal": {**store["sync"]["wallets"]["withdrawal"], **(parsed.get("wallets", {}).get("withdrawal", {}))},
                },
                "telegram": {**store["sync"]["telegram"], **parsed.get("telegram", {})},
            }
        if manual_row and manual_row["value"]:
            store["manualReconciliations"] = json.loads(manual_row["value"]) or []

        deposit_sync = ensure_wallet_sync_state(store["sync"], "deposit")
        withdrawal_sync = ensure_wallet_sync_state(store["sync"], "withdrawal")
        wallet_tracked_address = str(deposit_sync.get("trackedAddress") or "").lower()
        withdrawal_tracked_address = str(withdrawal_sync.get("trackedAddress") or "").lower()
        telegram_tracked_address = str(store["sync"]["telegram"].get("trackedAddress") or "").lower()
        if (
            (wallet_tracked_address and wallet_tracked_address != CONFIG["deposit_wallet_address"])
            or (withdrawal_tracked_address and withdrawal_tracked_address != CONFIG["withdrawal_wallet_address"])
            or (telegram_tracked_address and telegram_tracked_address != CONFIG["deposit_wallet_address"])
        ):
            store["sync"]["wallets"]["deposit"] = {
                **create_wallet_sync_state(),
                "trackedAddress": CONFIG["deposit_wallet_address"],
            }
            store["sync"]["wallets"]["withdrawal"] = {
                **create_wallet_sync_state(),
                "trackedAddress": CONFIG["withdrawal_wallet_address"],
            }
            store["sync"]["telegram"] = {
                **default_store()["sync"]["telegram"],
                "trackedAddress": CONFIG["deposit_wallet_address"],
            }

        store["sync"]["wallets"]["deposit"]["trackedAddress"] = CONFIG["deposit_wallet_address"]
        store["sync"]["wallets"]["withdrawal"]["trackedAddress"] = CONFIG["withdrawal_wallet_address"]
        store["sync"]["telegram"]["trackedAddress"] = CONFIG["deposit_wallet_address"]
        store["sync"]["telegram"].setdefault("importCursors", {})

        tx_rows = connection.execute(
            """
            SELECT
              tx_key,
              hash,
              wallet_address,
              block_number,
              sender,
              recipient,
              token_symbol,
              token_name,
              contract_address,
              amount,
              timestamp,
              raw_json
            FROM transactions
            ORDER BY datetime(timestamp) DESC, block_number DESC
            """
        ).fetchall()
        transactions = []
        for row in tx_rows:
            tx = {
                "txKey": row["tx_key"],
                "hash": row["hash"],
                "walletAddress": row["wallet_address"],
                "blockNumber": row["block_number"],
                "from": row["sender"],
                "to": row["recipient"],
                "tokenSymbol": row["token_symbol"],
                "tokenName": row["token_name"],
                "contractAddress": row["contract_address"],
                "amount": row["amount"],
                "timestamp": row["timestamp"],
                "raw": json.loads(row["raw_json"] or "{}"),
            }
            tx["txKey"] = tx["txKey"] or build_transaction_key(tx)
            if is_tracked_usdt_transaction(tx):
                transactions.append(tx)
        store["transactions"] = transactions

        message_rows = connection.execute(
            """
            SELECT
              id,
              message_id,
              chat_id,
              sender,
              amount,
              timestamp,
              text,
              raw_json
            FROM telegram_messages
            ORDER BY datetime(timestamp) DESC, id DESC
            """
        ).fetchall()
        messages: list[dict[str, Any]] = []
        for row in message_rows:
            raw = json.loads(row["raw_json"] or "{}")
            row_message = {
                "id": row["id"],
                "messageId": row["message_id"],
                "chatId": row["chat_id"],
                "sender": row["sender"],
                "amount": row["amount"],
                "timestamp": row["timestamp"],
                "text": row["text"],
                "raw": raw,
            }
            if message_references_current_wallet(row_message.get("text")) or message_references_current_wallet(raw.get("html")):
                upsert_telegram_message({"telegramMessages": messages}, normalize_imported_telegram_message(row_message))
        store["telegramMessages"] = sorted(messages, key=lambda item: item.get("timestamp") or "", reverse=True)
    return store


def write_store(store: dict[str, Any]) -> None:
    with connect_db() as connection:
        connection.execute("BEGIN")
        connection.execute("DELETE FROM transactions")
        connection.execute("DELETE FROM telegram_messages")

        for row in store.get("transactions", []):
            connection.execute(
                """
                INSERT INTO transactions (
                  tx_key, hash, wallet_address, block_number, sender, recipient, token_symbol,
                  token_name, contract_address, amount, timestamp, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("txKey") or build_transaction_key(row),
                    row.get("hash"),
                    row.get("walletAddress"),
                    row.get("blockNumber"),
                    row.get("from"),
                    row.get("to"),
                    row.get("tokenSymbol"),
                    row.get("tokenName"),
                    row.get("contractAddress"),
                    row.get("amount"),
                    row.get("timestamp"),
                    json.dumps(row.get("raw") or {}),
                ),
            )

        for row in store.get("telegramMessages", []):
            connection.execute(
                """
                INSERT INTO telegram_messages (
                  id, tx_hash, fees_tx_hash, message_id, chat_id, sender, amount, timestamp, text, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row.get("id"),
                    row.get("raw", {}).get("txHash"),
                    row.get("raw", {}).get("feesTxHash"),
                    row.get("messageId"),
                    row.get("chatId"),
                    row.get("sender"),
                    row.get("amount"),
                    row.get("timestamp"),
                    row.get("text"),
                    json.dumps(row.get("raw") or {}),
                ),
            )

        connection.execute(
            """
            INSERT INTO sync_state (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            ("store", json.dumps(store.get("sync") or default_store()["sync"])),
        )
        connection.execute(
            """
            INSERT INTO sync_state (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            ("manual_reconciliations", json.dumps(store.get("manualReconciliations") or [])),
        )
        connection.commit()


def amount_from_raw(value: str, decimals: int = 18) -> float:
    return int(value, 16 if str(value).startswith("0x") else 10) / (10 ** decimals)


def normalize_rpc_transaction(log: dict[str, Any], timestamp: str, wallet_address: str) -> dict[str, Any]:
    from_address = f"0x{log['topics'][1][-40:]}".lower()
    to_address = f"0x{log['topics'][2][-40:]}".lower()
    tx = {
        "id": str(log["transactionHash"]).lower(),
        "hash": str(log["transactionHash"]).lower(),
        "walletAddress": wallet_address,
        "blockNumber": int(log["blockNumber"], 16),
        "from": from_address,
        "to": to_address,
        "tokenSymbol": "USDT",
        "tokenName": "Tether USD",
        "contractAddress": str(log["address"]).lower(),
        "amount": int(log["data"], 16) / (10 ** 18),
        "timestamp": timestamp,
        "raw": {
            "walletAddress": wallet_address,
            "logIndex": int(log.get("logIndex", "0x0"), 16),
            "removed": log.get("removed", False),
            "transactionIndex": int(log.get("transactionIndex", "0x0"), 16),
        },
    }
    tx["txKey"] = build_transaction_key(tx)
    return tx


def normalize_bscscan_transaction(tx: dict[str, Any], wallet_address: str) -> dict[str, Any]:
    normalized = {
        "id": str(tx.get("hash") or "").lower(),
        "hash": str(tx.get("hash") or "").lower(),
        "walletAddress": wallet_address,
        "blockNumber": int(tx.get("blockNumber") or 0),
        "from": str(tx.get("from") or "").lower(),
        "to": str(tx.get("to") or "").lower(),
        "tokenSymbol": tx.get("tokenSymbol"),
        "tokenName": tx.get("tokenName"),
        "contractAddress": str(tx.get("contractAddress") or "").lower(),
        "amount": float(tx.get("value") or 0) / (10 ** int(tx.get("tokenDecimal") or 18)),
        "timestamp": datetime.fromtimestamp(int(tx.get("timeStamp") or 0), tz=timezone.utc).isoformat(),
        "raw": {
            **tx,
            "walletAddress": wallet_address,
        },
    }
    normalized["txKey"] = build_transaction_key(normalized)
    return normalized


def is_tracked_usdt_transaction(tx: dict[str, Any]) -> bool:
    return (
        str(tx.get("walletAddress") or "").lower() in {
            CONFIG["deposit_wallet_address"],
            CONFIG["withdrawal_wallet_address"],
        }
        and str(tx.get("contractAddress") or "").lower() == CONFIG["usdt_contract"]
    )


def extract_amount(text: str | None) -> float | None:
    if not text:
        return None
    patterns = [
        re.compile(r"(?:usdt|usd|amount|received|payment)[^\d]{0,10}(\d+(?:[.,]\d{1,6})?)", re.I),
        re.compile(r"(\d+(?:[.,]\d{1,6})?)\s*(?:usdt|usd)", re.I),
        re.compile(r"(?:rs|inr)[^\d]{0,10}(\d+(?:[.,]\d{1,2})?)", re.I),
    ]
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return float(match.group(1).replace(",", ""))
    return None


def extract_hash(text: str | None) -> str | None:
    if not text:
        return None
    match = re.search(r"0x[a-fA-F0-9]{64}", text)
    return match.group(0).lower() if match else None


def extract_hashes(text: str | None) -> list[str]:
    if not text:
        return []
    return [match.group(0).lower() for match in re.finditer(r"0x[a-fA-F0-9]{64}", text)]


def extract_labeled_hash(text: str | None, label: str) -> str | None:
    if not text:
        return None
    escaped = re.escape(label)
    match = re.search(rf"{escaped}\s*:?\s*(0x[a-fA-F0-9]{{64}})", text, re.I)
    return match.group(1).lower() if match else None


def normalize_imported_telegram_message(message: dict[str, Any]) -> dict[str, Any]:
    text = message.get("text") or ""
    raw = message.get("raw") or {}
    button_labels = raw.get("buttonLabels") if isinstance(raw.get("buttonLabels"), list) else []
    tx_hash = raw.get("txHash") or extract_labeled_hash(text, "Transaction Hash") or extract_hash(text)
    fees_tx_hash = raw.get("feesTxHash") or extract_labeled_hash(text, "Fees")
    return {
        "id": message.get("id") or str(uuid.uuid4()),
        "messageId": message.get("messageId"),
        "chatId": message.get("chatId") or "telegram-export",
        "sender": message.get("sender") or "unknown",
        "text": text,
        "amount": message.get("amount") if message.get("amount") is not None else extract_amount(text),
        "timestamp": message.get("timestamp") or datetime.now(timezone.utc).isoformat(),
        "raw": {
            **raw,
            "txHash": tx_hash,
            "feesTxHash": fees_tx_hash,
            "allHashes": raw.get("allHashes") or extract_hashes(text),
            "buttonLabels": button_labels,
            "hasViewTxButton": raw.get("hasViewTxButton", "View Tx" in button_labels),
            "hasViewFeesTxButton": raw.get("hasViewFeesTxButton", "View Fees Tx" in button_labels),
        },
    }


def find_telegram_message_index(store: dict[str, Any], message: dict[str, Any]) -> int:
    tx_hash = message.get("raw", {}).get("txHash")
    if tx_hash:
        for index, current in enumerate(store["telegramMessages"]):
            if current.get("raw", {}).get("txHash") == tx_hash:
                return index
    fees_tx_hash = message.get("raw", {}).get("feesTxHash")
    if fees_tx_hash:
        for index, current in enumerate(store["telegramMessages"]):
            if current.get("raw", {}).get("feesTxHash") == fees_tx_hash:
                return index
    if message.get("chatId") and message.get("messageId") is not None:
        for index, current in enumerate(store["telegramMessages"]):
            if current.get("chatId") == message.get("chatId") and current.get("messageId") == message.get("messageId"):
                return index
    for index, current in enumerate(store["telegramMessages"]):
        if current.get("id") == message.get("id"):
            return index
    return -1


def merge_telegram_message(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    return normalize_imported_telegram_message(
        {
            **existing,
            **incoming,
            "id": existing.get("id") or incoming.get("id"),
            "messageId": incoming.get("messageId", existing.get("messageId")),
            "chatId": incoming.get("chatId") or existing.get("chatId"),
            "sender": incoming.get("sender") or existing.get("sender"),
            "text": incoming.get("text") or existing.get("text"),
            "amount": incoming.get("amount") if incoming.get("amount") is not None else existing.get("amount"),
            "timestamp": incoming.get("timestamp") or existing.get("timestamp"),
            "raw": {
                **(existing.get("raw") or {}),
                **(incoming.get("raw") or {}),
                "buttonLabels": (incoming.get("raw", {}).get("buttonLabels") or existing.get("raw", {}).get("buttonLabels") or []),
                "hasViewTxButton": incoming.get("raw", {}).get("hasViewTxButton", existing.get("raw", {}).get("hasViewTxButton", False)),
                "hasViewFeesTxButton": incoming.get("raw", {}).get("hasViewFeesTxButton", existing.get("raw", {}).get("hasViewFeesTxButton", False)),
                "txHash": incoming.get("raw", {}).get("txHash") or existing.get("raw", {}).get("txHash"),
                "feesTxHash": incoming.get("raw", {}).get("feesTxHash") or existing.get("raw", {}).get("feesTxHash"),
                "allHashes": incoming.get("raw", {}).get("allHashes") or existing.get("raw", {}).get("allHashes") or [],
            },
        }
    )


def upsert_telegram_message(store: dict[str, Any], message: dict[str, Any]) -> bool:
    existing_index = find_telegram_message_index(store, message)
    if existing_index >= 0:
        store["telegramMessages"][existing_index] = merge_telegram_message(store["telegramMessages"][existing_index], message)
        return False
    store["telegramMessages"].append(message)
    return True


def normalize_cursor_chat_key(chat_id: str | None) -> str:
    return str(chat_id or "telegram-export")


def get_telegram_import_cursor(store: dict[str, Any], chat_id: str | None) -> dict[str, Any] | None:
    return store.get("sync", {}).get("telegram", {}).get("importCursors", {}).get(normalize_cursor_chat_key(chat_id))


def update_telegram_import_cursor(store: dict[str, Any], message: dict[str, Any]) -> None:
    store["sync"]["telegram"].setdefault("importCursors", {})
    key = normalize_cursor_chat_key(message.get("chatId"))
    existing = store["sync"]["telegram"]["importCursors"].get(key, {})
    existing_time = parse_any_datetime(existing.get("latestTimestamp")).timestamp() if existing.get("latestTimestamp") else 0
    message_time = parse_any_datetime(message.get("timestamp")).timestamp() if message.get("timestamp") else 0
    if message_time >= existing_time:
        store["sync"]["telegram"]["importCursors"][key] = {
            "latestTimestamp": message.get("timestamp") or existing.get("latestTimestamp"),
            "latestTxHash": message.get("raw", {}).get("txHash") or existing.get("latestTxHash"),
            "latestMessageId": message.get("messageId", existing.get("latestMessageId")),
            "updatedAt": datetime.now(timezone.utc).isoformat(),
        }


def should_skip_imported_telegram_message(store: dict[str, Any], message: dict[str, Any]) -> bool:
    cursor = get_telegram_import_cursor(store, message.get("chatId"))
    if not cursor or not cursor.get("latestTimestamp"):
        return False
    cursor_time = parse_any_datetime(cursor["latestTimestamp"]).timestamp()
    message_time = parse_any_datetime(message.get("timestamp")).timestamp() if message.get("timestamp") else 0
    if message_time > cursor_time:
        return False
    if message.get("raw", {}).get("txHash"):
        return find_telegram_message_index(store, message) >= 0
    if message.get("messageId") is not None:
        return message_time <= cursor_time
    return False


def is_primary_telegram_message(message: dict[str, Any]) -> bool:
    return bool(
        message.get("raw", {}).get("hasViewTxButton")
        and not message.get("raw", {}).get("hasViewFeesTxButton")
        and float(message.get("amount") or 0) >= CONFIG["min_transaction_amount"]
    )


def is_settlement_telegram_message(message: dict[str, Any]) -> bool:
    return bool(
        message.get("raw", {}).get("hasViewTxButton")
        and message.get("raw", {}).get("hasViewFeesTxButton")
        and float(message.get("amount") or 0) >= CONFIG["min_transaction_amount"]
    )


def is_qualifying_telegram_message(message: dict[str, Any]) -> bool:
    return is_primary_telegram_message(message) or is_settlement_telegram_message(message)


def find_settlement_message(message: dict[str, Any], settlement_messages: list[dict[str, Any]], settlement_by_hash: dict[str, dict[str, Any]], window_ms: int) -> dict[str, Any] | None:
    if not is_primary_telegram_message(message):
        return message if is_settlement_telegram_message(message) else None
    if message.get("raw", {}).get("txHash") and settlement_by_hash.get(message["raw"]["txHash"]):
        return settlement_by_hash[message["raw"]["txHash"]]
    message_time = parse_any_datetime(message["timestamp"]).timestamp() * 1000
    for candidate in settlement_messages:
        if candidate.get("amount") is None or message.get("amount") is None:
            continue
        candidate_time = parse_any_datetime(candidate["timestamp"]).timestamp() * 1000
        within_amount = abs(float(candidate["amount"]) - float(message["amount"])) <= CONFIG["amount_tolerance"]
        is_later = candidate_time >= message_time
        within_time = abs(candidate_time - message_time) <= window_ms
        if within_amount and is_later and within_time:
            return candidate
    return None


def get_message_net_amount(message: dict[str, Any]) -> float:
    amount = float(message.get("amount") or 0)
    fee_amount = float(message.get("raw", {}).get("feesAmount") or 0)
    return round(amount - fee_amount, 8)


def apply_manual_reconciliations(store: dict[str, Any], reconciled_transactions: list[dict[str, Any]], reconciled_messages: list[dict[str, Any]]) -> None:
    reconciliations = store.get("manualReconciliations") or []
    if not reconciliations:
        return
    transactions_by_hash = {tx["hash"]: tx for tx in reconciled_transactions}
    messages_by_id = {msg["id"]: msg for msg in reconciled_messages}
    for override in reconciliations:
        tx = transactions_by_hash.get(override.get("walletHash"))
        if not tx:
            continue
        messages = [messages_by_id[item] for item in override.get("telegramMessageIds", []) if item in messages_by_id]
        if not messages:
            continue
        gross_amount = round(sum(float(msg.get("amount") or 0) for msg in messages), 8)
        net_amount = round(sum(get_message_net_amount(msg) for msg in messages), 8)
        latest_timestamp = sorted([msg.get("timestamp") for msg in messages if msg.get("timestamp")])[-1] if messages else None
        summary_text = override.get("note") or f"Manual batch match ({len(messages)} Telegram messages)"

        tx["reconciliation"] = {
            **tx.get("reconciliation", {}),
            "status": "matched",
            "matchedMessageId": override.get("id") or "manual-batch",
            "matchedMessageText": summary_text,
            "matchedMessageAmount": gross_amount,
            "matchedMessageTimestamp": latest_timestamp,
            "manualBatch": True,
            "manualBatchMessageCount": len(messages),
            "manualBatchNetAmount": net_amount,
        }

        for message in messages:
            message["reconciliation"] = {
                **message.get("reconciliation", {}),
                "status": "matched",
                "matchedTransactionHash": tx["hash"],
                "matchedTransactionAmount": get_message_net_amount(message),
                "matchedTransactionTimestamp": tx.get("timestamp"),
                "matchedTransactionFrom": tx.get("from"),
                "landedInWallet": True,
                "manualBatch": True,
                "manualBatchMessageCount": len(messages),
                "manualBatchGrossAmount": gross_amount,
                "manualBatchNetAmount": net_amount,
            }


def reconcile(store: dict[str, Any]) -> dict[str, Any]:
    window_ms = CONFIG["match_window_minutes"] * 60 * 1000
    transactions = sorted(store.get("transactions", []), key=lambda item: item.get("timestamp") or "", reverse=True)
    all_messages = sorted(
        [msg for msg in store.get("telegramMessages", []) if is_qualifying_telegram_message(msg)],
        key=lambda item: item.get("timestamp") or "",
        reverse=True,
    )
    settlement_messages = [msg for msg in all_messages if is_settlement_telegram_message(msg)]
    primary_messages = [msg for msg in all_messages if is_primary_telegram_message(msg)] if any(is_primary_telegram_message(msg) for msg in all_messages) else settlement_messages
    used_hashes: set[str] = set()
    settlement_by_hash = {msg.get("raw", {}).get("txHash"): msg for msg in settlement_messages if msg.get("raw", {}).get("txHash")}

    reconciled_messages = []
    for message in primary_messages:
        settlement_message = find_settlement_message(message, settlement_messages, settlement_by_hash, window_ms)
        effective_message = settlement_message or message
        effective_time = parse_any_datetime(effective_message["timestamp"]).timestamp() * 1000
        direct_hash_match = None
        if effective_message.get("raw", {}).get("txHash"):
            for tx in transactions:
                if tx["hash"] == effective_message["raw"]["txHash"] and tx["hash"] not in used_hashes:
                    direct_hash_match = tx
                    break
        match = direct_hash_match
        if match is None:
            for tx in transactions:
                if tx["hash"] in used_hashes:
                    continue
                if effective_message.get("amount") is None or tx.get("amount") is None:
                    continue
                within_amount = abs(float(tx["amount"]) - float(effective_message["amount"])) <= CONFIG["amount_tolerance"]
                within_time = abs(parse_any_datetime(tx["timestamp"]).timestamp() * 1000 - effective_time) <= window_ms
                if within_amount and within_time:
                    match = tx
                    break
        if match:
            used_hashes.add(match["hash"])
        reconciled_messages.append(
            {
                **message,
                "reconciliation": {
                    "status": "matched" if match else "unmatched",
                    "linkedSettlementMessageId": settlement_message.get("id") if settlement_message else None,
                    "linkedSettlementTimestamp": settlement_message.get("timestamp") if settlement_message else None,
                    "linkedSettlementAmount": settlement_message.get("amount") if settlement_message else None,
                    "originTxHash": message.get("raw", {}).get("txHash"),
                    "originTxUrl": f"https://bscscan.com/tx/{message['raw']['txHash']}" if message.get("raw", {}).get("txHash") else None,
                    "settlementTxHash": effective_message.get("raw", {}).get("txHash"),
                    "settlementTxUrl": f"https://bscscan.com/tx/{effective_message['raw']['txHash']}" if effective_message.get("raw", {}).get("txHash") else None,
                    "feesTxHash": effective_message.get("raw", {}).get("feesTxHash"),
                    "feesTxUrl": f"https://bscscan.com/tx/{effective_message['raw']['feesTxHash']}" if effective_message.get("raw", {}).get("feesTxHash") else None,
                    "feesAmount": effective_message.get("raw", {}).get("feesAmount"),
                    "feesCurrency": effective_message.get("raw", {}).get("feesCurrency"),
                    "matchedTransactionHash": match.get("hash") if match else None,
                    "matchedTransactionAmount": match.get("amount") if match else None,
                    "matchedTransactionTimestamp": match.get("timestamp") if match else None,
                    "matchedTransactionFrom": match.get("from") if match else None,
                    "landedInWallet": bool(match),
                    "txHash": effective_message.get("raw", {}).get("txHash"),
                    "txUrl": f"https://bscscan.com/tx/{effective_message['raw']['txHash']}" if effective_message.get("raw", {}).get("txHash") else None,
                },
            }
        )

    matched_hashes = {msg["reconciliation"]["matchedTransactionHash"] for msg in reconciled_messages if msg["reconciliation"].get("matchedTransactionHash")}
    reconciled_transactions = []
    for tx in transactions:
        matched_message = next((msg for msg in reconciled_messages if msg["reconciliation"].get("matchedTransactionHash") == tx["hash"]), None)
        reconciled_transactions.append(
            {
                **tx,
                "reconciliation": {
                    "status": "matched" if tx["hash"] in matched_hashes else "unmatched",
                    "matchedMessageId": matched_message.get("id") if matched_message else None,
                    "matchedMessageText": matched_message.get("text") if matched_message else None,
                    "matchedMessageAmount": matched_message.get("amount") if matched_message else None,
                    "matchedMessageTimestamp": matched_message.get("timestamp") if matched_message else None,
                },
            }
        )
    apply_manual_reconciliations(store, reconciled_transactions, reconciled_messages)
    return {"transactions": reconciled_transactions, "telegramMessages": reconciled_messages}


def dedupe_transactions(transactions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result = []
    for tx in transactions:
        key = tx.get("txKey") or build_transaction_key(tx)
        if key in seen:
            continue
        seen.add(key)
        result.append(tx)
    return result


def message_references_current_wallet(text: str | None) -> bool:
    match = re.search(r"0x[a-fA-F0-9]{4,}\.\.\.[a-fA-F0-9]{4,}", str(text or ""))
    if not match:
        return True
    prefix, suffix = match.group(0).lower().split("...")
    return CONFIG["deposit_wallet_address"].startswith(prefix) and CONFIG["deposit_wallet_address"].endswith(suffix)


def decode_telegram_export_text(text: str | None) -> str:
    return (
        str(text or "")
        .replace("Ã°Å¸â€™Â°", "💰")
        .replace("Ã¢â‚¬Â¦", "…")
        .replace("Ã¢â‚¬â„¢", "'")
        .replace("Ã¢â‚¬Å“", '"')
        .replace("Ã¢â‚¬", '"')
        .replace("Ã‚", "")
        .strip()
    )


def parse_telegram_export_timestamp(value: str | None) -> str | None:
    match = re.match(r"^(\d{2})\.(\d{2})\.(\d{4}) (\d{2}):(\d{2}):(\d{2}) UTC([+-]\d{2}):(\d{2})$", str(value or ""))
    if not match:
        return None
    day, month, year, hour, minute, second, offset_hour, offset_minute = match.groups()
    iso_value = f"{year}-{month}-{day}T{hour}:{minute}:{second}{offset_hour}:{offset_minute}"
    try:
        return datetime.fromisoformat(iso_value).astimezone(timezone.utc).isoformat()
    except ValueError:
        return None


def parse_telegram_export_html(html: str, filename: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    chat_title = soup.select_one(".page_header .text.bold")
    chat_title_text = chat_title.get_text(strip=True) if chat_title else "Telegram Export"
    messages = []
    for node in soup.select(".message.default, .message.default.clearfix"):
        body = node.select_one(".body")
        if body is None:
            continue
        message_id_raw = (node.get("id") or "").replace("message", "")
        date_node = body.select_one(".pull_right.date.details")
        timestamp_title = date_node.get("title") if date_node else ""
        from_name_node = body.select_one(".from_name")
        text_node = body.select_one(".text")
        if not timestamp_title or text_node is None:
            continue
        html_text = str(text_node)
        button_labels = [item.get_text(strip=True) for item in body.select(".bot_buttons_table .bot_button div")]
        plain_text = decode_telegram_export_text(
            re.sub(r"\s{2,}", " ", text_node.get_text("\n", strip=True).replace(" \n", "\n").replace("\n ", "\n")).strip()
        )
        if not plain_text:
            continue
        normalized_timestamp = parse_telegram_export_timestamp(timestamp_title)
        if not normalized_timestamp:
            continue
        amount = extract_amount(plain_text)
        tx_hash = extract_labeled_hash(plain_text, "Transaction Hash") or extract_hash(plain_text)
        fees_tx_hash = extract_labeled_hash(plain_text, "Fees")
        messages.append(
            {
                "id": f"export-{filename}-{message_id_raw or uuid.uuid4()}",
                "messageId": int(message_id_raw) if message_id_raw.isdigit() else None,
                "chatId": chat_title_text,
                "sender": from_name_node.get_text(strip=True) if from_name_node else chat_title_text,
                "text": plain_text,
                "amount": amount,
                "timestamp": normalized_timestamp,
                "raw": {
                    "source": "telegram-export-html",
                    "filename": filename,
                    "txHash": tx_hash,
                    "feesTxHash": fees_tx_hash,
                    "allHashes": extract_hashes(plain_text),
                    "buttonLabels": button_labels,
                    "hasViewTxButton": "View Tx" in button_labels,
                    "hasViewFeesTxButton": "View Fees Tx" in button_labels,
                    "html": html_text,
                },
            }
        )
    return messages


def parse_any_datetime(value: str | None) -> datetime:
    if not value:
        return datetime.fromtimestamp(0, tz=timezone.utc)
    normalized = value.replace("Z", "+00:00")
    parsed = datetime.fromisoformat(normalized)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


async def rpc_call(method: str, params: list[Any]) -> Any:
    payload = {"jsonrpc": "2.0", "id": 1, "method": method, "params": params}
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(CONFIG["rpc_url"], json=payload)
        response.raise_for_status()
        data = response.json()
    if data.get("error"):
        raise RuntimeError(data["error"].get("message") or f"{method} failed")
    return data.get("result")


def is_rpc_limit_error(error: Exception) -> bool:
    message = str(error).lower()
    return any(
        token in message
        for token in [
            "limit exceeded",
            "query returned more than",
            "response size exceeded",
            "block range is too wide",
            "rate limit",
            "too many results",
        ]
    )


async def rpc_call_with_backoff(method: str, params: list[Any], attempts: int = 5) -> Any:
    delay = 0.4
    last_error: Exception | None = None
    endpoints = [CONFIG["rpc_url"], *CONFIG["rpc_fallback_urls"]]
    for _ in range(attempts):
        for endpoint in endpoints:
            try:
                payload = {"jsonrpc": "2.0", "id": 1, "method": method, "params": params}
                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.post(endpoint, json=payload)
                    response.raise_for_status()
                    data = response.json()
                if data.get("error"):
                    raise RuntimeError(data["error"].get("message") or f"{method} failed")
                return data.get("result")
            except Exception as exc:
                last_error = exc
                if not is_rpc_limit_error(exc):
                    continue
        await asyncio.sleep(delay)
        delay *= 1.8
    if last_error is not None:
        raise last_error
    raise RuntimeError(f"{method} failed")


def to_hex_quantity(value: int) -> str:
    return hex(max(0, int(value)))


def padded_topic_address(address: str) -> str:
    return f"0x{'0' * 24}{address.lower().replace('0x', '')}"


async def get_wallet_snapshot(address: str, label: str) -> dict[str, Any]:
    balance_of_selector = "0x70a08231"
    padded_owner = f"{address.lower().replace('0x', ''):0>64}"
    balance_call = balance_of_selector + padded_owner
    native_balance, token_balance, transaction_count, latest_block = await asyncio.gather(
        rpc_call("eth_getBalance", [address, "latest"]),
        rpc_call("eth_call", [{"to": CONFIG["usdt_contract"], "data": balance_call}, "latest"]),
        rpc_call("eth_getTransactionCount", [address, "latest"]),
        rpc_call("eth_blockNumber", []),
    )
    return {
        "label": label,
        "address": address,
        "network": "BNB Smart Chain",
        "bscscanAddressUrl": f"https://bscscan.com/address/{address}",
        "bscscanTokenUrl": f"https://bscscan.com/token/{CONFIG['usdt_contract']}?a={address}",
        "nativeSymbol": "BNB",
        "tokenSymbol": "USDT",
        "nativeBalance": int(native_balance, 16) / (10 ** 18),
        "tokenBalance": int(token_balance, 16) / (10 ** 18),
        "transactionCount": int(transaction_count, 16),
        "latestBlock": int(latest_block, 16),
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "error": None,
    }


async def resolve_fee_transfer(fees_tx_hash: str) -> dict[str, Any] | None:
    tx, receipt = await asyncio.gather(
        rpc_call("eth_getTransactionByHash", [fees_tx_hash]),
        rpc_call("eth_getTransactionReceipt", [fees_tx_hash]),
    )
    if not tx or not receipt:
        return None
    transfer_logs = [log for log in receipt.get("logs", []) if log.get("topics", [None])[0] == TRANSFER_TOPIC]
    preferred_log = next((log for log in transfer_logs if str(log.get("address") or "").lower() == CONFIG["usdt_contract"]), transfer_logs[0] if transfer_logs else None)
    if preferred_log:
        contract_address = str(preferred_log.get("address") or "").lower()
        currency = "USDT" if contract_address == CONFIG["usdt_contract"] else contract_address
        return {
            "amount": int(preferred_log.get("data") or "0x0", 16) / (10 ** 18),
            "currency": currency,
            "contractAddress": contract_address,
        }
    tx_input = str(tx.get("input") or "")
    if tx_input.startswith(TRANSFER_SELECTOR) and str(tx.get("to") or "").lower() == CONFIG["usdt_contract"]:
        encoded_value = tx_input[74:138]
        return {
            "amount": int(encoded_value or "0", 16) / (10 ** 18),
            "currency": "USDT",
            "contractAddress": CONFIG["usdt_contract"],
        }
    if int(tx.get("value") or "0x0", 16) > 0:
        return {
            "amount": int(tx["value"], 16) / (10 ** 18),
            "currency": "BNB",
            "contractAddress": None,
        }
    return None


async def enrich_telegram_fees(store: dict[str, Any]) -> dict[str, Any]:
    targets = [
        message
        for message in store.get("telegramMessages", [])
        if is_qualifying_telegram_message(message)
        and message.get("raw", {}).get("feesTxHash")
        and message.get("raw", {}).get("feesAmount") is None
    ]
    if not targets:
        return store
    updated = False
    cache: dict[str, dict[str, Any] | None] = {}
    for message in targets:
        fees_tx_hash = message.get("raw", {}).get("feesTxHash")
        if not fees_tx_hash:
            continue
        try:
            if fees_tx_hash not in cache:
                cache[fees_tx_hash] = await resolve_fee_transfer(fees_tx_hash)
            fee_details = cache[fees_tx_hash]
            if fee_details is not None:
                message["raw"]["feesAmount"] = fee_details["amount"]
                message["raw"]["feesCurrency"] = fee_details["currency"]
                message["raw"]["feesContractAddress"] = fee_details["contractAddress"]
                updated = True
        except Exception as exc:
            message["raw"]["feesResolveError"] = str(exc)
    if updated:
        write_store(store)
    return store


async def fetch_transactions_from_rpc(wallet_address: str, start_block: int, full_history: bool = False) -> list[dict[str, Any]]:
    latest_block = int(await rpc_call("eth_blockNumber", []), 16)
    initial_from_block = start_block if start_block > 0 else (0 if full_history else max(0, latest_block - 12000))
    recipient_topic = padded_topic_address(wallet_address)
    sender_topic = padded_topic_address(wallet_address)
    logs: list[dict[str, Any]] = []

    async def fetch_window(from_block: int, to_block: int) -> list[dict[str, Any]]:
        incoming_batch = await rpc_call_with_backoff(
            "eth_getLogs",
            [{
                "address": CONFIG["usdt_contract"],
                "fromBlock": to_hex_quantity(from_block),
                "toBlock": to_hex_quantity(to_block),
                "topics": [TRANSFER_TOPIC, None, recipient_topic],
            }],
        )
        outgoing_batch = await rpc_call_with_backoff(
            "eth_getLogs",
            [{
                "address": CONFIG["usdt_contract"],
                "fromBlock": to_hex_quantity(from_block),
                "toBlock": to_hex_quantity(to_block),
                "topics": [TRANSFER_TOPIC, sender_topic, None],
            }],
        )
        return [*(incoming_batch or []), *(outgoing_batch or [])]

    max_step = 120 if full_history else 240
    min_step = 1
    current_block = initial_from_block
    while current_block <= latest_block:
        step = min(max_step, latest_block - current_block)
        while True:
            to_block = min(latest_block, current_block + step)
            try:
                logs.extend(await fetch_window(current_block, to_block))
                current_block = to_block + 1
                if step < max_step:
                    max_step = min(max_step, step * 2)
                break
            except Exception as exc:
                if not is_rpc_limit_error(exc) or step <= min_step:
                    raise
                step = max(min_step, step // 2)
                await asyncio.sleep(0.25)

    unique_block_numbers = sorted({int(log["blockNumber"], 16) for log in logs})
    blocks_by_number: dict[int, Any] = {}
    for block_number in unique_block_numbers:
        blocks_by_number[block_number] = await rpc_call("eth_getBlockByNumber", [to_hex_quantity(block_number), False])
    transactions = []
    for log in logs:
        block = blocks_by_number.get(int(log["blockNumber"], 16)) or {}
        timestamp = datetime.fromtimestamp(int(block.get("timestamp", "0x0"), 16), tz=timezone.utc).isoformat()
        transactions.append(normalize_rpc_transaction(log, timestamp, wallet_address))
    return transactions


async def fetch_transactions_from_bscscan_api(wallet_address: str, full_history: bool = False) -> list[dict[str, Any]]:
    transactions: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    max_pages = 1000 if full_history else 5
    async with httpx.AsyncClient(timeout=30.0) as client:
        for page in range(1, max_pages + 1):
            response = await client.get(
                CONFIG["bsc_api_base"],
                params={
                    "module": "account",
                    "action": "tokentx",
                    "address": wallet_address,
                    "contractaddress": CONFIG["usdt_contract"],
                    "page": page,
                    "offset": 100,
                    "sort": "desc",
                    "apikey": CONFIG["bsc_api_key"],
                },
            )
            response.raise_for_status()
            payload = response.json() or {}
            result = payload.get("result")
            if not isinstance(result, list):
                raise RuntimeError(f"BscScan API returned {payload.get('message') or 'an invalid response'} for {wallet_address}")
            if payload.get("status") == "0" and payload.get("message") not in {None, "No transactions found"}:
                raise RuntimeError(f"BscScan API returned {payload.get('message')} for {wallet_address}")
            if not result:
                break
            page_rows = [normalize_bscscan_transaction(item, wallet_address) for item in result]
            new_rows = []
            for row in page_rows:
                key = row.get("txKey") or build_transaction_key(row)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                new_rows.append(row)
            if not new_rows:
                break
            transactions.extend(new_rows)
            if len(result) < 100:
                break
    return transactions


def count_bscscan_transfer_rows(html: str) -> int:
    soup = BeautifulSoup(html, "html.parser")
    tbody = soup.find("tbody")
    if tbody is None:
        return 0
    return len(tbody.find_all("tr", recursive=False))


def extract_address_from_href(href: str | None) -> str | None:
    match = re.search(r"/address/(0x[a-fA-F0-9]{40})", str(href or ""))
    return match.group(1).lower() if match else None


def extract_address_from_cell(cell) -> str | None:
    link = cell.find("a", href=re.compile(r"/address/")) if cell else None
    href_address = extract_address_from_href(link.get("href") if link else None)
    if href_address:
        return href_address
    for attr in ("data-clipboard-text", "data-address", "data-highlight-target"):
        node = cell.find(attrs={attr: True}) if cell else None
        if node:
            direct_match = re.search(r"0x[a-fA-F0-9]{40}", str(node.get(attr) or ""))
            if direct_match:
                return direct_match.group(0).lower()
    text_match = re.search(r"0x[a-fA-F0-9]{40}", cell.get_text(" ", strip=True) if cell else "")
    return text_match.group(0).lower() if text_match else None


def parse_bscscan_transfer_page(html: str, wallet_address: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    rows: list[dict[str, Any]] = []
    normalized_wallet_address = str(wallet_address or "").lower()
    tbody = soup.find("tbody")
    if tbody is None:
        return rows
    for row in tbody.find_all("tr", recursive=False):
        cells = row.find_all("td", recursive=False)
        if len(cells) < 13:
            continue
        hash_link = cells[1].find("a", href=re.compile(r"^/tx/"))
        hash_value = hash_link.get_text(strip=True).lower() if hash_link else ""
        block_text = cells[4].get_text(strip=True)
        timestamp_text = cells[5].get_text(strip=True)
        from_cell = cells[8]
        direction_text = cells[9].get_text(strip=True).upper()
        to_cell = cells[10]
        to_text = to_cell.get_text(strip=True)
        amount_span = cells[11].find(class_="td_showAmount")
        amount_text = amount_span.get_text(strip=True) if amount_span else cells[11].get_text(strip=True)
        token_link = cells[12].find("a", href=re.compile(r"^/token/"))
        token_href = (token_link.get("href") or "").lower() if token_link else ""
        if not hash_value or CONFIG["usdt_contract"] not in token_href:
            continue
        from_address = extract_address_from_cell(from_cell)
        to_address = extract_address_from_cell(to_cell)
        if direction_text == "IN":
            from_address = from_address or from_cell.get_text(" ", strip=True).lower()
            to_address = normalized_wallet_address
        elif direction_text == "OUT":
            from_address = normalized_wallet_address
            to_address = to_address or to_text.lower()
        else:
            from_address = from_address or from_cell.get_text(" ", strip=True).lower()
            to_address = to_address or to_text.lower() or normalized_wallet_address
        normalized_to = str(to_address or "").lower()
        normalized_from = str(from_address or "").lower()
        if normalized_to != normalized_wallet_address and normalized_from != normalized_wallet_address:
            continue
        amount = float(amount_text.replace(",", "")) if amount_text else 0.0
        timestamp = datetime.strptime(f"{timestamp_text} UTC", "%Y-%m-%d %H:%M:%S %Z").replace(tzinfo=timezone.utc).isoformat() if timestamp_text else datetime.now(timezone.utc).isoformat()
        normalized = {
            "id": hash_value,
            "hash": hash_value,
            "walletAddress": normalized_wallet_address,
            "blockNumber": int(block_text or 0),
            "from": normalized_from,
            "to": normalized_to,
            "tokenSymbol": "USDT",
            "tokenName": "Tether USD",
            "contractAddress": CONFIG["usdt_contract"],
            "amount": amount,
            "timestamp": timestamp,
            "raw": {
                "source": "bscscan-html",
            },
        }
        normalized["txKey"] = build_transaction_key(normalized)
        rows.append(normalized)
    return rows


async def fetch_transactions_from_bscscan_html(wallet_address: str, full_history: bool = False) -> list[dict[str, Any]]:
    transactions: list[dict[str, Any]] = []
    seen_keys: set[str] = set()
    max_pages = 1000 if full_history else 5
    async with httpx.AsyncClient(timeout=30.0, headers={"User-Agent": "Mozilla/5.0"}) as client:
        for page in range(1, max_pages + 1):
            response = await client.get(
                "https://bscscan.com/tokentxns",
                params={"a": wallet_address, "p": page, "ps": 100},
            )
            response.raise_for_status()
            html = response.text
            raw_row_count = count_bscscan_transfer_rows(html)
            page_rows = parse_bscscan_transfer_page(html, wallet_address)
            if not page_rows:
                break
            new_rows = []
            for row in page_rows:
                key = row.get("txKey") or build_transaction_key(row)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                new_rows.append(row)
            if not new_rows:
                break
            transactions.extend(new_rows)
            if raw_row_count < 100:
                break
    return transactions


async def sync_wallet_transactions(store: dict[str, Any], wallet_key: str, mode: str = "incremental") -> None:
    wallet_address = get_wallet_address(wallet_key)
    wallet_sync = ensure_wallet_sync_state(store["sync"], wallet_key)
    sync_time = datetime.now(timezone.utc).isoformat()
    sync_source = "rpc"
    try:
        incoming = await fetch_transactions_from_rpc(wallet_address, 0 if mode == "full" else wallet_sync.get("lastBlock") or 0, full_history=(mode == "full"))
    except Exception as primary_error:
        try:
            incoming = await fetch_transactions_from_bscscan_api(wallet_address, full_history=(mode == "full"))
            sync_source = "bscscan-api"
        except Exception as api_error:
            try:
                incoming = await fetch_transactions_from_bscscan_html(wallet_address, full_history=(mode == "full"))
                sync_source = "bscscan-html"
            except Exception as fallback_error:
                raise RuntimeError(f"{wallet_key} wallet sync failed. Primary: {primary_error}. API: {api_error}. HTML: {fallback_error}") from fallback_error
    incoming = sorted(
        [tx for tx in incoming if float(tx.get("amount") or 0) >= CONFIG["min_transaction_amount"]],
        key=lambda item: item.get("timestamp") or "",
        reverse=True,
    )
    other_transactions = [tx for tx in store.get("transactions", []) if tx.get("walletAddress") != wallet_address]

    if mode == "full":
        store["transactions"] = dedupe_transactions(other_transactions + incoming)
        wallet_sync["lastBlock"] = max([int(tx.get("blockNumber") or 0) for tx in incoming], default=0)
        wallet_sync["lastCheckedAt"] = sync_time
        wallet_sync["lastError"] = None
        wallet_sync["lastSource"] = sync_source
        wallet_sync["baselineTimestamp"] = incoming[-1]["timestamp"] if incoming else None
        wallet_sync["initialized"] = True
        wallet_sync["lastMode"] = "full"
        wallet_sync["lastFullSyncAt"] = sync_time
        return

    known_keys = {
        tx.get("txKey") or build_transaction_key(tx)
        for tx in store.get("transactions", [])
        if tx.get("walletAddress") == wallet_address
    }

    if not wallet_sync.get("initialized"):
        newest_incoming_time = parse_any_datetime(incoming[0]["timestamp"]).timestamp() if incoming else datetime.now(timezone.utc).timestamp()
        initial_baseline = newest_incoming_time - CONFIG["initial_backfill_hours"] * 60 * 60
        initial_transactions = [tx for tx in incoming if parse_any_datetime(tx["timestamp"]).timestamp() >= initial_baseline]
        store["transactions"] = list(other_transactions)
        for tx in initial_transactions:
            key = tx.get("txKey") or build_transaction_key(tx)
            if key not in known_keys:
                store["transactions"].append(tx)
                known_keys.add(key)
        wallet_sync["lastBlock"] = max([int(tx.get("blockNumber") or 0) for tx in incoming], default=wallet_sync.get("lastBlock") or 0)
        wallet_sync["lastCheckedAt"] = sync_time
        wallet_sync["lastError"] = None
        wallet_sync["lastSource"] = sync_source
        wallet_sync["baselineTimestamp"] = datetime.fromtimestamp(initial_baseline, tz=timezone.utc).isoformat()
        wallet_sync["initialized"] = True
        wallet_sync["lastMode"] = "incremental"
        wallet_sync["lastIncrementalSyncAt"] = sync_time
        store["transactions"] = sorted(store["transactions"], key=lambda item: item.get("timestamp") or "", reverse=True)
        return

    baseline_timestamp = parse_any_datetime(wallet_sync.get("baselineTimestamp")).timestamp() if wallet_sync.get("baselineTimestamp") else 0
    filtered_incoming = []
    for tx in incoming:
        tx_time = parse_any_datetime(tx["timestamp"]).timestamp()
        if tx_time >= baseline_timestamp or int(tx.get("blockNumber") or 0) > int(wallet_sync.get("lastBlock") or 0):
            filtered_incoming.append(tx)

    existing_wallet_transactions = [tx for tx in store.get("transactions", []) if tx.get("walletAddress") == wallet_address]
    store["transactions"] = other_transactions + existing_wallet_transactions
    for tx in filtered_incoming:
        key = tx.get("txKey") or build_transaction_key(tx)
        if key not in known_keys:
            store["transactions"].append(tx)
            known_keys.add(key)
        wallet_sync["lastBlock"] = max(int(wallet_sync.get("lastBlock") or 0), int(tx.get("blockNumber") or 0))
    store["transactions"] = sorted(store["transactions"], key=lambda item: item.get("timestamp") or "", reverse=True)
    wallet_sync["lastCheckedAt"] = sync_time
    wallet_sync["lastError"] = None
    wallet_sync["lastSource"] = sync_source
    wallet_sync["lastMode"] = "incremental"
    wallet_sync["lastIncrementalSyncAt"] = sync_time


async def sync_transactions_safe(mode: str = "incremental") -> dict[str, Any]:
    store = read_store()
    try:
        await sync_wallet_transactions(store, "deposit", mode)
        if CONFIG["withdrawal_wallet_address"] != CONFIG["deposit_wallet_address"]:
            await sync_wallet_transactions(store, "withdrawal", mode)
        else:
            withdrawal_sync = ensure_wallet_sync_state(store["sync"], "withdrawal")
            withdrawal_sync["trackedAddress"] = CONFIG["withdrawal_wallet_address"]
            withdrawal_sync["lastCheckedAt"] = store["sync"]["wallets"]["deposit"].get("lastCheckedAt")
            withdrawal_sync["lastError"] = "Withdrawal wallet is identical to deposit wallet. Set WITHDRAWAL_WALLET_ADDRESS to a different address."
        write_store(store)
        return store
    except Exception as exc:
        timestamp = datetime.now(timezone.utc).isoformat()
        ensure_wallet_sync_state(store["sync"], "deposit")["lastCheckedAt"] = timestamp
        ensure_wallet_sync_state(store["sync"], "deposit")["lastError"] = str(exc)
        ensure_wallet_sync_state(store["sync"], "withdrawal")["lastCheckedAt"] = timestamp
        ensure_wallet_sync_state(store["sync"], "withdrawal")["lastError"] = str(exc)
        write_store(store)
        return store


async def run_audit_sync(mode: str = "incremental") -> dict[str, Any]:
    async with audit_sync_lock:
        return await sync_transactions_safe(mode)


def schedule_audit_sync(mode: str = "incremental") -> bool:
    global audit_sync_task
    if audit_sync_task and not audit_sync_task.done():
        return False
    audit_sync_task = asyncio.create_task(run_audit_sync(mode))
    return True


def audit_sync_needed(state: dict[str, Any]) -> bool:
    if state["depositTransactions"] or state["withdrawalTransactions"] or state["interWalletTransfers"]:
        last_checked = state["sync"]["wallets"]["deposit"].get("lastCheckedAt")
        if last_checked:
            age = (datetime.now(timezone.utc) - parse_any_datetime(last_checked)).total_seconds()
            return age >= max(30, CONFIG["poll_interval_ms"] / 1000)
    return True


async def get_state() -> dict[str, Any]:
    store = read_store()
    await enrich_telegram_fees(store)
    usdt_transactions = [tx for tx in store.get("transactions", []) if is_tracked_usdt_transaction(tx)]
    deposit_transactions = sorted(
        [tx for tx in usdt_transactions if tx.get("walletAddress") == CONFIG["deposit_wallet_address"]],
        key=lambda item: item.get("timestamp") or "",
        reverse=True,
    )
    withdrawal_transactions = sorted(
        [tx for tx in usdt_transactions if tx.get("walletAddress") == CONFIG["withdrawal_wallet_address"]],
        key=lambda item: item.get("timestamp") or "",
        reverse=True,
    )
    deposit_incoming_transactions = [tx for tx in deposit_transactions if tx.get("to") == CONFIG["deposit_wallet_address"]]
    reconciled = reconcile({**store, "transactions": deposit_incoming_transactions})
    inter_wallet_transfers = dedupe_transactions(
        [
            tx
            for tx in usdt_transactions
            if (
                (tx.get("from") == CONFIG["deposit_wallet_address"] and tx.get("to") == CONFIG["withdrawal_wallet_address"])
                or (tx.get("from") == CONFIG["withdrawal_wallet_address"] and tx.get("to") == CONFIG["deposit_wallet_address"])
            )
        ]
    )
    inter_wallet_transfers = sorted(inter_wallet_transfers, key=lambda item: item.get("timestamp") or "", reverse=True)

    wallet_snapshots = {}
    for wallet_key, label in {"deposit": "Deposit wallet", "withdrawal": "Withdrawal wallet"}.items():
        address = get_wallet_address(wallet_key)
        try:
            wallet_snapshots[wallet_key] = await get_wallet_snapshot(address, label)
        except Exception as exc:
            wallet_snapshots[wallet_key] = {
                "label": label,
                "address": address,
                "network": "BNB Smart Chain",
                "bscscanAddressUrl": f"https://bscscan.com/address/{address}",
                "bscscanTokenUrl": f"https://bscscan.com/token/{CONFIG['usdt_contract']}?a={address}",
                "nativeSymbol": "BNB",
                "tokenSymbol": "USDT",
                "nativeBalance": None,
                "tokenBalance": None,
                "transactionCount": None,
                "latestBlock": None,
                "updatedAt": datetime.now(timezone.utc).isoformat(),
                "error": str(exc),
            }

    return {
        "config": {
            "depositWalletAddress": CONFIG["deposit_wallet_address"],
            "withdrawalWalletAddress": CONFIG["withdrawal_wallet_address"],
            "walletsAreDistinct": CONFIG["deposit_wallet_address"] != CONFIG["withdrawal_wallet_address"],
            "usdtContract": CONFIG["usdt_contract"],
            "pollIntervalMs": CONFIG["poll_interval_ms"],
            "rpcUrl": CONFIG["rpc_url"],
            "telegramChatId": CONFIG["telegram_chat_id"] or None,
            "hasBscApiKey": bool(CONFIG["bsc_api_key"]),
            "hasTelegramBotToken": bool(CONFIG["telegram_bot_token"]),
            "matchWindowMinutes": CONFIG["match_window_minutes"],
            "amountTolerance": CONFIG["amount_tolerance"],
            "initialBackfillHours": CONFIG["initial_backfill_hours"],
            "minTransactionAmount": CONFIG["min_transaction_amount"],
        },
        "walletSnapshots": wallet_snapshots,
        "sync": store["sync"],
        "depositTransactions": reconciled["transactions"],
        "withdrawalTransactions": withdrawal_transactions,
        "interWalletTransfers": inter_wallet_transfers,
        "telegramMessages": reconciled["telegramMessages"],
    }


def build_workbook(state: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    telegram_sheet = workbook.create_sheet("Telegram")
    transactions_sheet = workbook.create_sheet("Deposit Transactions")

    matched_messages = [msg for msg in state["telegramMessages"] if msg.get("reconciliation", {}).get("status") == "matched"]
    unmatched_messages = [msg for msg in state["telegramMessages"] if msg.get("reconciliation", {}).get("status") != "matched"]

    def sum_amounts(items: list[dict[str, Any]], selector) -> float:
        return sum(float(selector(item) or 0) for item in items)

    matched_rate = round((len(matched_messages) / len(state["telegramMessages"]) * 100), 2) if state["telegramMessages"] else 0
    matched_variance = sum(
        abs(float(msg.get("amount") or 0) - float(msg.get("reconciliation", {}).get("matchedTransactionAmount") or 0))
        for msg in matched_messages
    )
    summary_rows = [
        ("Deposit wallet", CONFIG["deposit_wallet_address"]),
        ("Withdrawal wallet", CONFIG["withdrawal_wallet_address"]),
        ("Telegram-first match base", "Enabled"),
        ("Deposit wallet transactions", len(state["depositTransactions"])),
        ("Withdrawal wallet transactions", len(state["withdrawalTransactions"])),
        ("Inter-wallet transfers", len(state["interWalletTransfers"])),
        ("Telegram messages", len(state["telegramMessages"])),
        ("Matched telegram messages", len(matched_messages)),
        ("Unmatched telegram messages", len(unmatched_messages)),
        ("Telegram total amount", sum_amounts(state["telegramMessages"], lambda msg: msg.get("amount"))),
        ("Matched wallet amount", sum_amounts(matched_messages, lambda msg: msg.get("reconciliation", {}).get("matchedTransactionAmount"))),
        ("Unmatched telegram amount", sum_amounts(unmatched_messages, lambda msg: msg.get("amount"))),
        ("Matched rate %", matched_rate),
        ("Matched variance amount", matched_variance),
        ("Deposit wallet total amount", sum_amounts(state["depositTransactions"], lambda tx: tx.get("amount"))),
        ("Withdrawal wallet total amount", sum_amounts(state["withdrawalTransactions"], lambda tx: tx.get("amount"))),
        ("Last deposit wallet sync", state["sync"]["wallets"]["deposit"].get("lastCheckedAt") or ""),
        ("Last deposit wallet error", state["sync"]["wallets"]["deposit"].get("lastError") or ""),
        ("Last withdrawal wallet sync", state["sync"]["wallets"]["withdrawal"].get("lastCheckedAt") or ""),
        ("Last withdrawal wallet error", state["sync"]["wallets"]["withdrawal"].get("lastError") or ""),
        ("Last telegram sync", state["sync"]["telegram"].get("lastCheckedAt") or ""),
        ("Last telegram error", state["sync"]["telegram"].get("lastError") or ""),
    ]
    summary_sheet.append(["metric", "value"])
    for row in summary_rows:
        summary_sheet.append(list(row))

    telegram_sheet.append(
        [
            "id",
            "chatId",
            "sender",
            "amount",
            "timestamp",
            "text",
            "status",
            "landedInWallet",
            "originTxHash",
            "originTxUrl",
            "settlementTxHash",
            "settlementTxUrl",
            "feesTxHash",
            "feesTxUrl",
            "matchedWalletHash",
            "matchedWalletUrl",
            "matchedWalletAmount",
            "matchedWalletTimestamp",
            "linkedSettlementTimestamp",
        ]
    )
    for msg in state["telegramMessages"]:
        reconciliation = msg.get("reconciliation", {})
        telegram_sheet.append(
            [
                msg.get("id"),
                msg.get("chatId"),
                msg.get("sender"),
                msg.get("amount"),
                msg.get("timestamp"),
                msg.get("text"),
                reconciliation.get("status") or "unmatched",
                "Yes" if reconciliation.get("landedInWallet") else "No",
                reconciliation.get("originTxHash") or "",
                reconciliation.get("originTxUrl") or "",
                reconciliation.get("settlementTxHash") or "",
                reconciliation.get("settlementTxUrl") or "",
                reconciliation.get("feesTxHash") or "",
                reconciliation.get("feesTxUrl") or "",
                reconciliation.get("matchedTransactionHash") or "",
                f"https://bscscan.com/tx/{reconciliation['matchedTransactionHash']}" if reconciliation.get("matchedTransactionHash") else "",
                reconciliation.get("matchedTransactionAmount") or "",
                reconciliation.get("matchedTransactionTimestamp") or "",
                reconciliation.get("linkedSettlementTimestamp") or "",
            ]
        )

    transactions_sheet.append(
        ["hash", "txUrl", "amount", "timestamp", "from", "to", "status", "matchedMessageAmount", "matchedMessageTimestamp", "matchedMessageText"]
    )
    for tx in state["depositTransactions"]:
        reconciliation = tx.get("reconciliation", {})
        transactions_sheet.append(
            [
                tx.get("hash"),
                f"https://bscscan.com/tx/{tx['hash']}" if tx.get("hash") else "",
                tx.get("amount"),
                tx.get("timestamp"),
                tx.get("from"),
                tx.get("to"),
                reconciliation.get("status"),
                reconciliation.get("matchedMessageAmount"),
                reconciliation.get("matchedMessageTimestamp"),
                reconciliation.get("matchedMessageText"),
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@router.get("/audit-api/state", include_in_schema=False)
async def audit_state():
    state = await get_state()
    if audit_sync_needed(state):
        schedule_audit_sync("incremental")
    return state


@router.post("/audit-api/sync/wallet", include_in_schema=False)
async def audit_sync_wallet(payload: dict[str, Any]):
    mode = "full" if payload.get("mode") == "full" else "incremental"
    scheduled = schedule_audit_sync(mode)
    state = await get_state()
    return JSONResponse(status_code=202, content={"scheduled": scheduled, "mode": mode, "state": state})


@router.post("/audit-api/sync/telegram", include_in_schema=False)
async def audit_sync_telegram(payload: dict[str, Any]):
    mode = "all" if payload.get("mode") == "all" else "incremental"
    store = read_store()
    store["sync"]["telegram"]["lastCheckedAt"] = datetime.now(timezone.utc).isoformat()
    store["sync"]["telegram"]["lastError"] = (
        None
        if store["sync"]["telegram"].get("historicalImportSupported")
        else "Telegram Bot API cannot backfill old chat history; it only receives new messages after the bot is added and running."
    )
    store["sync"]["telegram"]["lastMode"] = mode
    write_store(store)
    return await get_state()


@router.post("/audit-api/telegram/import", include_in_schema=False)
async def audit_import_telegram(files: list[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")
    store = read_store()
    imported = 0
    skipped = 0
    for file in files:
        html = (await file.read()).decode("utf-8", errors="ignore")
        parsed = [normalize_imported_telegram_message(item) for item in parse_telegram_export_html(html, file.filename or "upload.html")]
        for message in parsed:
            if should_skip_imported_telegram_message(store, message):
                skipped += 1
                continue
            if upsert_telegram_message(store, message):
                imported += 1
            update_telegram_import_cursor(store, message)
    store["telegramMessages"] = sorted(store["telegramMessages"], key=lambda item: item.get("timestamp") or "", reverse=True)
    store["sync"]["telegram"]["lastCheckedAt"] = datetime.now(timezone.utc).isoformat()
    store["sync"]["telegram"]["lastError"] = None
    store["sync"]["telegram"]["lastMode"] = "import"
    store["sync"]["telegram"]["lastMessageAt"] = store["telegramMessages"][0]["timestamp"] if store["telegramMessages"] else store["sync"]["telegram"].get("lastMessageAt")
    write_store(store)
    return {
        "imported": imported,
        "skipped": skipped,
        "totalMessages": len(store["telegramMessages"]),
        "state": await get_state(),
    }


@router.get("/audit-api/export.xlsx", include_in_schema=False)
async def audit_export():
    workbook = build_workbook(await get_state())
    headers = {"Content-Disposition": "attachment; filename=usdt-dashboard-export.xlsx"}
    return Response(
        content=workbook.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/audit-api/status", include_in_schema=False)
async def audit_status():
    return {
        "available": True,
        "reason": "Python-native audit API is active.",
        "processRunning": True,
        "target": "fastapi-native",
        "database": str(AUDIT_DB_FILE),
    }
